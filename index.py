from openpyxl import Workbook
import pandas as pd
from numpy import random
from datetime import datetime
import re


def excelToText(filename):
    excelFile = pd.read_excel("./data/" + filename + ".xlsx")
    excelFile.to_csv("./text/" + filename + ".txt", index=None, header=True)


def prepareJob(df):
    header = ("JobKey", "JobId", "JobTitle",
              "JobLevel", "JobTag")

    book = Workbook()
    sheet = book.active
    sheet.append(header)

    keys = []
    jobKey = 1
    for index, row in df.iterrows():
        jobTitle = str(row[3])
        jobLevel = str(row[2])
        jobTag = "None" if (pd.isna(row[8])) else str(row[8])
        jobId = re.sub('\W', '', jobTitle.strip().upper(
        ) + jobLevel.strip().upper() + jobTag.strip().upper())

        data = (
            jobKey,
            jobId,
            jobTitle,
            jobLevel,
            jobTag,
        )

        if (jobId not in keys):
            print(data)
            keys.append(jobId)
            sheet.append(data)

            jobKey += 1

    book.save("data/job.xlsx")
    excelToText("job")
    print("Done")


def prepareCompany(df):
    header = ("CompanyKey", "CompanyId", "CompanyName",
              "CityId", "City", "State", "Country")

    book = Workbook()
    sheet = book.active
    sheet.append(header)

    keys = []
    companyKey = 1
    for index, row in df.iterrows():
        companyName = str(row[1]).upper()
        cityId = row[14]
        city = ""
        state = ""
        country = ""

        location = row[5]

        # City
        if (location.count(",") > 0):
            temp = location.split(",")
            city = temp[0].strip()

        # State
        if (location.count(",") > 0):
            temp = location.split(",")
            state = temp[1].strip()

        # Country
        if (location.count(",") == 1):
            country = "United States"
        elif (location.count(",") == 2):
            temp = location.split(",")
            country = temp[2].strip()

        companyId = re.sub(
            '\W', '', companyName.strip().upper() + str(cityId).strip())

        data = (
            companyKey,
            companyId,
            companyName,
            cityId,
            city,
            state,
            country,
        )

        if (companyId not in keys):
            keys.append(companyId)
            print(data)
            sheet.append(data)

            companyKey += 1

    book.save("data/company.xlsx")
    excelToText("company")
    print("Done")


def prepareEducation(df):
    header = ("EducationKey", "EducationId",
              "EducationName", "EducationNameAbbr")

    book = Workbook()
    sheet = book.active
    sheet.append(header)

    keys = []
    educationKey = 1
    for index, row in df.iterrows():
        educationId = "NOT APPLICABLE" if pd.isna(row[28]) else str(row[28])

        if (educationId == "NOT APPLICABLE"):
            educationName = "NOT APPLICABLE"
            educationNameAbbr = "NOT APPLICABLE"
        elif (educationId == "PhD"):
            educationName = "Doctorate Degree"
            educationNameAbbr = "PhD"
        elif (educationId == "Master's Degree"):
            educationName = "Master's Degree"
            educationNameAbbr = "MS"
        elif (educationId == "Bachelor's Degree"):
            educationName = "Bachelor's Degree"
            educationNameAbbr = "BS"
        elif (educationId == "Some College"):
            educationName = "Some College"
            educationNameAbbr = "SC"
        else:
            educationName = "Highschool"
            educationNameAbbr = "HS"

        data = (
            educationKey,
            educationId,
            educationName,
            educationNameAbbr,
        )

        if (educationId not in keys):
            keys.append(educationId)
            print(data)
            sheet.append(data)

            educationKey += 1

    book.save("data/education.xlsx")
    excelToText("education")
    print("Done")


def prepareDemographic(df):
    header = ("DemographicKey", "DemographicId",
              "Gender", "GenderAbbr", "Race", )

    book = Workbook()
    sheet = book.active
    sheet.append(header)

    keys = []
    demographicKey = 1
    for index, row in df.iterrows():
        gender = random.choice(["Male", "Female"]) if pd.isna(
            row[12]) or row[12] == "Title: Senior Software Engineer" else str(row[12])
        genderAbbr = gender[0]
        race = "White" if pd.isna(row[27]) else str(row[27])
        demographicId = re.sub('\W', '', race.strip().upper() +
                               gender.strip().upper())

        data = (
            demographicKey,
            demographicId,
            gender,
            genderAbbr,
            race,
        )

        if (demographicId not in keys):
            keys.append(demographicId)
            print(data)
            sheet.append(data)

            demographicKey += 1

    book.save("data/demographic.xlsx")
    excelToText("demographic")
    print("Done")


def prepareDate(df):
    header = ("DateKey",  "CalendarDate", "CalendarDateChar", "Year", "QuarterNumber", "QuarterName", "MonthNumber",
              "MonthName", "MonthNameAbbr", "DayOfYear", "DayOfMonth", "DayOfWeek", "DayName", "DayNameAbbr")

    book = Workbook()
    sheet = book.active
    sheet.append(header)

    keys = []
    dateKey = 1
    for index, row in df.iterrows():
        calendarDate = datetime.strptime(
            row[0], '%m/%d/%Y %H:%M').date()
        calendarDateChar = str(calendarDate)
        year = calendarDate.year

        monthNumber = calendarDate.strftime("%m")
        monthName = calendarDate.strftime("%B")
        monthNameAbbr = calendarDate.strftime("%b")

        quarterNumber = 0
        quarterName = ""
        if int(monthNumber) >= 1 and int(monthNumber) <= 3:
            quarterNumber = 1
            quarterName = "Q1"
        elif int(monthNumber) >= 4 and int(monthNumber) <= 6:
            quarterNumber = 2
            quarterName = "Q2"
        elif int(monthNumber) >= 7 and int(monthNumber) <= 9:
            quarterNumber = 3
            quarterName = "Q3"
        elif int(monthNumber) >= 10 and int(monthNumber) <= 12:
            quarterNumber = 4
            quarterName = "Q4"

        dayOfYear = calendarDate.strftime("%j")
        dayOfMonth = calendarDate.strftime("%d")
        dayOfWeek = int(calendarDate.strftime("%w")) + 1
        dayName = calendarDate.strftime("%A")
        dayNameAbbr = calendarDate.strftime("%a")

        data = (
            dateKey,
            calendarDate,
            calendarDateChar,
            year,
            quarterNumber,
            quarterName,
            monthNumber,
            monthName,
            monthNameAbbr,
            dayOfYear,
            dayOfMonth,
            dayOfWeek,
            dayName,
            dayNameAbbr,
        )

        if (calendarDateChar not in keys):
            keys.append(calendarDateChar)
            print(data)
            sheet.append(data)

            dateKey += 1

    book.save("data/date.xlsx")
    excelToText("date")
    print("Done")


def prepareSalary(df):
    header = ("JobKey", "CompanyKey", "EducationKey", "DemographicKey", "DateKey",
              "BaseSalary", "TotalYearlyCompensation", "Bonus")

    book = Workbook()
    sheet = book.active
    sheet.append(header)

    keys = []
    for index, row in df.iterrows():
        jobTitle = str(row[3])
        jobLevel = str(row[2])
        jobTag = "None" if (pd.isna(row[8])) else str(row[8])
        jobKey = re.sub('\W', '', jobTitle.strip().upper(
        ) + jobLevel.strip().upper() + jobTag.strip().upper())

        companyName = str(row[1]).upper()
        cityId = row[14]
        companyKey = re.sub(
            '\W', '', companyName.strip().upper() + str(cityId).strip())

        educationKey = "NOT APPLICABLE" if pd.isna(row[28]) else str(row[28])

        baseSalary = 1000 if pd.isna(row[9]) or row[9] == 0 else row[9]
        totalYearlyCompensation = row[4]
        bonus = row[11]

        gender = random.choice(["Male", "Female"]) if pd.isna(
            row[12]) or row[12] == "Title: Senior Software Engineer" else str(row[12])
        race = "White" if pd.isna(row[27]) else str(row[27])
        demographicKey = re.sub('\W', '', race.strip().upper() +
                                gender.strip().upper())

        calendarDate = datetime.strptime(
            row[0], '%m/%d/%Y %H:%M').date()
        dateKey = str(calendarDate)

        data = (
            jobKey,
            companyKey,
            educationKey,
            demographicKey,
            dateKey,
            baseSalary,
            totalYearlyCompensation,
            bonus,
        )

        print(data)
        sheet.append(data)

    book.save("data/salary.xlsx")
    excelToText("salary")
    print("Done")


def main():
    df = pd.read_csv("source/source_data.csv", index_col=None)

    prepareJob(df)
    prepareCompany(df)
    prepareEducation(df)
    prepareDemographic(df)
    prepareDate(df)
    prepareSalary(df)


main()
